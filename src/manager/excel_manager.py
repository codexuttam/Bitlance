import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

class ExcelManager:
    def __init__(self, file_path="data/ceo_data.xlsx"):
        self.file_path = file_path
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

    def _parse_revenue(self, rev_str):
        """Helper to convert revenue string to float for sorting."""
        try:
            # Extract numbers and multipliers (e.g., "$500.5 billion" -> 500.5)
            # Match digits and decimals
            match = re.search(r'[\d\.,]+', str(rev_str).replace(',', ''))
            if match:
                return float(match.group())
        except:
            pass
        return 0.0

    def save_leads(self, df):
        try:
            # 1. Prepare Data
            required_cols = [
                "Full Name", "Company Name", "Industry", "Country", 
                "Email Address", "Mobile / Contact", "LinkedIn URL", 
                "Net Worth (USD)", "Company Revenue", "Data Source URL"
            ]
            
            for col in required_cols:
                if col not in df.columns:
                    df[col] = "N/A"

            # Add sort column
            df['_rev_numeric'] = df['Company Revenue'].apply(self._parse_revenue)
            df = df.sort_values(by='_rev_numeric', ascending=False)
            
            df_master = df[required_cols].copy()
            
            # Sheet 2: Email Ready (filtered)
            # We use 'Is_Email_Valid' if it exists, otherwise check if Email != "Contact Pending"
            if 'Is_Email_Valid' in df.columns:
                df_email_ready = df[df['Is_Email_Valid'] == True][required_cols].copy()
            else:
                df_email_ready = df[df['Email Address'].str.contains('@')][required_cols].copy()

            # 2. Save with professional formatting
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df_master.to_excel(writer, index=False, sheet_name='CEO Master List')
                df_email_ready.to_excel(writer, index=False, sheet_name='Email Ready')
                
                workbook = writer.book
                
                # Apply styles to both sheets
                for sheet_name in ['CEO Master List', 'Email Ready']:
                    ws = workbook[sheet_name]
                    
                    # Header Style
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

                    # Freeze Top Row
                    ws.freeze_panes = "A2"

                    # Auto-adjust column width
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[column].width = min(max_length + 2, 50)

                    # 3. Conditional Formatting for Emails
                    # Column E is 'Email Address' (5th column)
                    email_col_letter = 'E'
                    # Green for verified (contains @ and not placeholder)
                    # Yellow for unverified
                    
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    # Apply to email column cells (skip header)
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f"{email_col_letter}{row_idx}"]
                        email_val = str(cell.value)
                        if "@" in email_val and "corporate.com" not in email_val:
                            cell.fill = green_fill
                        elif "@" in email_val:
                            cell.fill = yellow_fill

            print(f"✅ Leads saved successfully to {self.file_path}")
            return True
        except Exception as e:
            print(f"❌ Error saving leads to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    def load_leads(self):
        if os.path.exists(self.file_path):
            return pd.read_excel(self.file_path)
        return None
